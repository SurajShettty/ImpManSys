"""Client implementation-plan exports (Excel + PDF), triggered by staff from
the Client Detail page. Unlike the client-portal schemas, these include full
internal detail (owner, internal notes, delay reasons) since distribution is
controlled by whoever downloads and sends the file, same trust model as the
existing client-list CSV export (routers/clients.py:export_clients).

PDF colours mirror the web app's palette (frontend/src/index.css) so the
export feels like the same product, not a generic report.
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app import models

# ---------------------------------------------------------------------------
# Palette (matches frontend/src/index.css :root)
# ---------------------------------------------------------------------------
PRIMARY = "#E31E24"
PRIMARY_DARK = "#C41A20"
PRIMARY_LIGHT = "#FDE8E8"
TEXT_HEADING = "#212529"
TEXT_LABEL = "#495057"
TEXT_MUTED = "#6c757d"
BORDER = "#dee2e6"
CARD_BG = "#fafafa"
PROGRESS_FILL = "#198754"
PROGRESS_TRACK = "#e9ecef"

TONE_COLORS = {
    "grey": ("#e9ecef", "#495057"),
    "blue": ("#cfe2ff", "#084298"),
    "amber": ("#fff3cd", "#664d03"),
    "green": ("#d1e7dd", "#0f5132"),
    "red": ("#f8d7da", "#842029"),
    "theme": ("#FCE5E5", "#9A1419"),
}
STATUS_TONE = {
    "Not Started": "grey", "In Progress": "blue", "On Hold": "amber",
    "Completed": "green", "Cancelled": "red",
    "Waiting for Client": "amber", "Waiting for Internal Team": "amber",
    "Blocked": "red", "Under Testing": "theme",
}
PRIORITY_TONE = {"Critical": "red", "High": "amber", "Medium": "blue", "Low": "grey"}

CONTENT_WIDTH = 18 * cm

SUMMARY_LABELS = [
    "Institution Type", "Region", "State", "Priority", "Status",
    "Implementation State", "CSM", "RM", "PM", "Sales Owner",
    "Contract Start", "Contract End", "Go-Live Date", "Kickoff Meeting Date",
    "Billing/Go-Live Date", "Master Data Status",
]


def _summary_values(client: models.Client) -> list:
    return [
        client.institution_type or "—",
        client.region or "—",
        client.state or "—",
        client.priority,
        client.status,
        client.implementation_state or "—",
        ", ".join(u.name for u in client.csms) or "—",
        ", ".join(u.name for u in client.rms) or "—",
        client.pm.name if client.pm else "—",
        client.sales_owner or "—",
        client.contract_start or "—",
        client.contract_end or "—",
        client.go_live_date or "—",
        client.kickoff_meeting_date or "—",
        client.billing_date or "—",
        client.master_data_status or "—",
    ]


def _active_phases(client: models.Client) -> list[models.Phase]:
    return [p for p in client.phases if not p.is_deleted]


def _active_modules(phase: models.Phase) -> list[models.PhaseModule]:
    return [pm for pm in phase.phase_modules if not pm.is_deleted]


def _active_activities(phase_module: models.PhaseModule) -> list[models.Activity]:
    return [a for a in phase_module.activities if not a.is_deleted]


PLAN_COLUMNS = [
    "Phase", "Module", "Activity", "Priority", "Status", "Progress %",
    "Start Date", "Due Date", "Owner", "Client SPOC", "UAT Proposed",
    "Client Response", "Delay Reason",
]


def _plan_row(phase: models.Phase, module: models.Module | None, activity: models.Activity) -> list:
    return [
        phase.name,
        module.name if module else "",
        activity.title,
        activity.priority,
        activity.status,
        activity.progress,
        activity.start_date or "",
        activity.due_date or "",
        activity.owner.name if activity.owner else "",
        activity.client_spoc or "",
        "Yes" if activity.uat_proposed else "No",
        activity.client_response or "",
        activity.delay_reason or "",
    ]


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def build_client_plan_workbook(client: models.Client) -> bytes:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor=PRIMARY_DARK.lstrip("#"))
    header_font = Font(bold=True, color="FFFFFF")

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Client", client.name])
    summary["A1"].font = header_font
    summary["A1"].fill = header_fill
    summary["B1"].fill = header_fill
    summary["B1"].font = Font(bold=True, color="FFFFFF")
    for label, value in zip(SUMMARY_LABELS, _summary_values(client)):
        summary.append([label, value])
    for row in summary.iter_rows(min_row=2, max_col=1):
        row[0].font = Font(bold=True)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 40

    plan = wb.create_sheet("Implementation Plan")
    plan.append(PLAN_COLUMNS)
    for cell in plan[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    plan.freeze_panes = "A2"

    for phase in _active_phases(client):
        for phase_module in _active_modules(phase):
            for activity in _active_activities(phase_module):
                plan.append(_plan_row(phase, phase_module.module, activity))

    widths = [20, 18, 30, 10, 16, 11, 12, 12, 18, 20, 12, 30, 30]
    for idx, width in enumerate(widths, start=1):
        plan.column_dimensions[plan.cell(row=1, column=idx).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF - small building blocks
# ---------------------------------------------------------------------------
def _pill(text: str, tone: str, max_width: float) -> Table:
    """A rounded, coloured badge. Text wraps (via Paragraph) instead of
    overflowing when it's too long for max_width - e.g. "Waiting for
    Internal Team" - so the pill always stays within its allotted column."""
    bg, fg = TONE_COLORS.get(tone, TONE_COLORS["grey"])
    pill_style = ParagraphStyle(
        "pill", fontName="Helvetica-Bold", fontSize=6.5, leading=7.5,
        textColor=colors.HexColor(fg), alignment=1,  # 1 = center
    )
    t = Table([[Paragraph(text, pill_style)]], colWidths=[max_width])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg)),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("ROUNDEDCORNERS", [5, 5, 5, 5]),
    ]))
    return t


def _status_pill(status: str, max_width: float) -> Table:
    return _pill(status, STATUS_TONE.get(status, "grey"), max_width)


def _priority_pill(priority: str, max_width: float) -> Table:
    return _pill(priority, PRIORITY_TONE.get(priority, "grey"), max_width)


def _progress_cell(value: float | None, total_width: float = 2.6 * cm) -> Table:
    """A mini bar + percentage label sized to total_width exactly, so it
    always fits the column it's placed in (bar + label proportions are
    derived from total_width rather than hardcoded, to avoid overflow)."""
    pct = max(0.0, min(100.0, value or 0.0))
    label_width = 0.85 * cm
    gap = 0.1 * cm
    bar_width = total_width - label_width - gap
    bar_height = 0.2 * cm
    filled = round(bar_width * pct / 100.0, 2)
    if filled <= 0.02:
        bar = Table([[""]], colWidths=[bar_width], rowHeights=[bar_height])
        bar.setStyle(TableStyle([("BACKGROUND", (0, 0), (0, 0), colors.HexColor(PROGRESS_TRACK)), ("ROUNDEDCORNERS", [2, 2, 2, 2])]))
    elif filled >= bar_width - 0.02:
        bar = Table([[""]], colWidths=[bar_width], rowHeights=[bar_height])
        bar.setStyle(TableStyle([("BACKGROUND", (0, 0), (0, 0), colors.HexColor(PROGRESS_FILL)), ("ROUNDEDCORNERS", [2, 2, 2, 2])]))
    else:
        bar = Table([["", ""]], colWidths=[filled, bar_width - filled], rowHeights=[bar_height])
        bar.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(PROGRESS_FILL)),
            ("BACKGROUND", (1, 0), (1, 0), colors.HexColor(PROGRESS_TRACK)),
            ("ROUNDEDCORNERS", [2, 2, 2, 2]),
        ]))
    label = Paragraph(f"{pct:.0f}%", ParagraphStyle("pct", fontName="Helvetica", fontSize=7, textColor=colors.HexColor(TEXT_MUTED)))
    cell = Table([[bar, label]], colWidths=[bar_width + gap, label_width])
    cell.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    return cell


def _page_decorator(client_name: str):
    def _decorate(canvas_obj, doc):
        canvas_obj.saveState()
        width, height = A4

        canvas_obj.setFillColor(colors.HexColor(PRIMARY))
        canvas_obj.rect(0, height - 0.3 * cm, width, 0.3 * cm, stroke=0, fill=1)

        canvas_obj.setFillColor(colors.HexColor(TEXT_MUTED))
        canvas_obj.setFont("Helvetica", 7.5)
        canvas_obj.drawString(1.5 * cm, height - 1.1 * cm, "DIGII  ·  IMPLEMENTATION MANAGEMENT SYSTEM")
        canvas_obj.drawRightString(width - 1.5 * cm, height - 1.1 * cm, date.today().strftime("%d %b %Y"))

        canvas_obj.setFillColor(colors.HexColor(TEXT_HEADING))
        canvas_obj.setFont("Helvetica-Bold", 14)
        canvas_obj.drawString(1.5 * cm, height - 1.68 * cm, f"{client_name} — Implementation Plan")

        canvas_obj.setStrokeColor(colors.HexColor(BORDER))
        canvas_obj.setLineWidth(0.6)
        canvas_obj.line(1.5 * cm, height - 1.9 * cm, width - 1.5 * cm, height - 1.9 * cm)

        canvas_obj.line(1.5 * cm, 1.3 * cm, width - 1.5 * cm, 1.3 * cm)
        canvas_obj.setFillColor(colors.HexColor(TEXT_MUTED))
        canvas_obj.setFont("Helvetica", 7.5)
        canvas_obj.drawString(1.5 * cm, 0.9 * cm, "Digii Implementation Management System")
        canvas_obj.drawRightString(width - 1.5 * cm, 0.9 * cm, f"Page {canvas_obj.getPageNumber()}")

        canvas_obj.restoreState()

    return _decorate


def _styles():
    base = getSampleStyleSheet()
    return {
        "snapshot_heading": ParagraphStyle("snapshot_heading", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=9, textColor=colors.HexColor(PRIMARY), spaceAfter=6),
        "kv_label": ParagraphStyle("kv_label", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=7.5, textColor=colors.HexColor(TEXT_LABEL), leading=10),
        "kv_value": ParagraphStyle("kv_value", parent=base["Normal"], fontName="Helvetica", fontSize=8.5, textColor=colors.HexColor(TEXT_HEADING), leading=11),
        "phase_name": ParagraphStyle("phase_name", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=11.5, textColor=colors.HexColor(TEXT_HEADING)),
        "module_heading": ParagraphStyle("module_heading", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=9.5, textColor=colors.HexColor(PRIMARY_DARK), spaceBefore=8, spaceAfter=4),
        "activity_cell": ParagraphStyle("activity_cell", parent=base["Normal"], fontName="Helvetica", fontSize=8, textColor=colors.HexColor(TEXT_HEADING), leading=10),
        "due_overdue": ParagraphStyle("due_overdue", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=8, textColor=colors.HexColor("#842029"), leading=10),
        "due_normal": ParagraphStyle("due_normal", parent=base["Normal"], fontName="Helvetica", fontSize=8, textColor=colors.HexColor(TEXT_MUTED), leading=10),
        "muted_italic": ParagraphStyle("muted_italic", parent=base["Normal"], fontName="Helvetica-Oblique", fontSize=8.5, textColor=colors.HexColor(TEXT_MUTED)),
    }


def _summary_card(client: models.Client, styles: dict) -> Table:
    pairs = list(zip(SUMMARY_LABELS, [str(v) for v in _summary_values(client)]))
    rows = []
    for i in range(0, len(pairs), 2):
        left = pairs[i]
        right = pairs[i + 1] if i + 1 < len(pairs) else ("", "")
        rows.append([
            Paragraph(left[0].upper(), styles["kv_label"]), Paragraph(left[1], styles["kv_value"]),
            Paragraph(right[0].upper(), styles["kv_label"]), Paragraph(right[1], styles["kv_value"]),
        ])
    col = [3.3 * cm, 5.7 * cm, 3.3 * cm, 5.7 * cm]
    table = Table(rows, colWidths=col)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(CARD_BG)),
        ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor(BORDER)),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 10), ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return table


def _phase_header(phase: models.Phase, styles: dict) -> Table:
    # Column budgets below already subtract this row's own padding (8pt each
    # side = ~0.28cm) so the pill/progress-bar flowables never exceed the
    # width available inside their cell.
    row = [
        Paragraph(phase.name, styles["phase_name"]),
        _status_pill(phase.status, max_width=1.6 * cm),
        _progress_cell(phase.progress, total_width=2.5 * cm),
    ]
    table = Table([row], colWidths=[10.5 * cm, 2.5 * cm, 3.4 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(PRIMARY_LIGHT)),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "CENTER"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    return table


def _activity_table(activities: list[models.Activity], styles: dict) -> Table:
    # Column widths already leave room for this table's own cell padding
    # (6pt each side = ~0.42cm) plus a safety margin, so pills/progress bars
    # never render wider than the column that contains them.
    today = date.today()
    header = ["Activity", "Priority", "Status", "Progress", "Due Date"]
    rows = [header]
    for a in activities:
        overdue = bool(a.due_date and a.due_date < today and a.status not in ("Completed", "Cancelled"))
        due_style = styles["due_overdue"] if overdue else styles["due_normal"]
        due_text = f"{a.due_date} (overdue)" if overdue and a.due_date else str(a.due_date or "—")
        rows.append([
            Paragraph(a.title, styles["activity_cell"]),
            _priority_pill(a.priority, max_width=1.5 * cm),
            _status_pill(a.status, max_width=2.0 * cm),
            _progress_cell(a.progress, total_width=2.5 * cm),
            Paragraph(due_text, due_style),
        ])

    table = Table(rows, colWidths=[7.0 * cm, 2.3 * cm, 2.8 * cm, 3.3 * cm, 2.5 * cm], repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY_DARK)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (1, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(BORDER)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    for i in range(1, len(rows)):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f8f9fa")))
    table.setStyle(TableStyle(style))
    return table


def build_client_plan_pdf(client: models.Client) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=2.6 * cm, bottomMargin=1.8 * cm,
        title=f"{client.name} - Implementation Plan",
    )
    styles = _styles()
    story = [
        Paragraph("CLIENT SNAPSHOT", styles["snapshot_heading"]),
        _summary_card(client, styles),
        Spacer(1, 0.7 * cm),
    ]

    phases = _active_phases(client)
    if not phases:
        story.append(Paragraph("No phases yet.", styles["muted_italic"]))

    for phase in phases:
        story.append(KeepTogether([_phase_header(phase, styles), Spacer(1, 0.2 * cm)]))
        modules = _active_modules(phase)
        if not modules:
            story.append(Paragraph("No modules added yet.", styles["muted_italic"]))
            story.append(Spacer(1, 0.4 * cm))
            continue

        for phase_module in modules:
            module_name = phase_module.module.name if phase_module.module else "Module"
            story.append(Paragraph(f"{module_name}  ·  {phase_module.progress:.0f}% complete", styles["module_heading"]))
            activities = _active_activities(phase_module)
            if not activities:
                story.append(Paragraph("No activities yet.", styles["muted_italic"]))
                continue
            story.append(_activity_table(activities, styles))

        story.append(Spacer(1, 0.5 * cm))

    doc.build(story, onFirstPage=_page_decorator(client.name), onLaterPages=_page_decorator(client.name))
    return buffer.getvalue()
